from bs4 import BeautifulSoup
import requests
import urllib.parse
import openpyxl
import logging
import json
import time
from pathlib import Path
from jsonpath_ng import parse
from requests.exceptions import RequestException


logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

c_handler = logging.StreamHandler()
c_handler.setLevel(logging.DEBUG)
c_format = logging.Formatter('%(asctime)s - %(levelname)s: %(message)s', datefmt='%H:%M:%S')
c_handler.setFormatter(c_format)
logger.addHandler(c_handler)

f_handler = logging.FileHandler('data_log.log')
f_handler.setLevel(logging.WARNING)
f_format = logging.Formatter('%(asctime)s - %(levelname)s: %(message)s', datefmt='%d-%m-%Y %H:%M:%S')
f_handler.setFormatter(f_format)
logger.addHandler(f_handler)

address = 'https://lubimyczytac.pl/'
file_path = Path('ebook.xlsx')
wb = openpyxl.load_workbook(file_path)
sheet = wb.active


def main(row_from, row_to):
    logger.info('Getting Started...\n')

    lc_headers = {'X-Csrf-Token': 'b42ee05dc19f593fb108c70c11a17e22', 'X-Requested-With': 'XMLHttpRequest'}
    counter = 1
    missing_items = []
    for i in range(row_from, row_to):
        title = sheet.cell(row=i + 1, column=2)
        author = sheet.cell(row=i + 1, column=1).value
        surname = author.split(' ', 1)[1] if author.find(' ') > 0 else author
        logger.info(f'Search item: {surname} - {title.value}.')
        uri1 = urllib.parse.quote(title.value + ' ' + surname)
        uri2 = urllib.parse.quote(title.value)

        url = f'https://lubimyczytac.pl/searcher/getsuggestions?phrase={uri1}'
        url2 = f'https://lubimyczytac.pl/searcher/getsuggestions?phrase={uri2}'

        def book_url(url):
            try:
                content = requests.get(url, headers=lc_headers).text
                soup = BeautifulSoup(content, 'html.parser')
                json_text = soup.get_text()
                json_data = json.loads(json_text)
                jsonpath_expression = parse('$.items.books.results[0].url')
                match = jsonpath_expression.find(json_data)
                return match
            except RequestException as e:
                logger.warning(f'Request error: {str(e)}')
                return None

        try:
            book_match = book_url(url)
            title_url = book_match[0].value
            logger.info(f'Loading the website {title_url}.')
        except:
            book_match = book_url(url2)
            if book_match is None or len(book_match) == 0:
                logger.warning(f'No book {title.value} found.')
                missing_items.append(title.value)
                counter += 1
                continue
            title_url = book_match[0].value
            logger.info(f'Loading the website {title_url}.')

        content2 = requests.get(title_url).text
        soup = BeautifulSoup(content2, 'html.parser')
        description = soup.find("div", {"class": "collapse-content"}).find('p')
        logger.info(f'Row {i}. {title.value} - saving the following data:')

        rate = soup.find('span', {'class': 'big-number'}).text.strip()
        rate = rate.replace(',', '.')
        sheet.cell(row=i + 1, column=4).value = float(rate)
        sheet.cell(row=i + 1, column=4).number_format = '#,##0.00'

        voters = soup.find('a', class_='btn-link').text.strip().split()[0]
        if voters.isdigit():
            sheet.cell(row=i + 1, column=5).value = int(voters)
        else:
            sheet.cell(row=i + 1, column=5).value = voters
        logger.info(f'   - rating {rate}, {voters} votes;')

        date_of_edition_soup = soup.select_one('dt:-soup-contains("Data wydania:")')
        date_of_edition = date_of_edition_soup.find_next_sibling('dd').text.strip() if date_of_edition_soup else ""
        sheet.cell(row=i + 1, column=8).value = date_of_edition

        date_of_1edition_soup = soup.select_one('dt:-soup-contains("Data 1. wydania:")')
        date_of_1edition = date_of_1edition_soup.find_next_sibling('dd').text.strip() if date_of_1edition_soup else ""
        sheet.cell(row=i + 1, column=7).value = date_of_1edition

        pages_no_soup = soup.select_one('dt:-soup-contains("Liczba stron:")')
        pages_no = pages_no_soup.find_next_sibling('dd').text.strip() if pages_no_soup else ""
        sheet.cell(row=i + 1, column=9).value = int(pages_no) if pages_no else ""

        category_soup = soup.select_one('dt:-soup-contains("Kategoria:")')
        category = category_soup.find_next_sibling('dd').text.strip() if category_soup else ""
        sheet.cell(row=i + 1, column=6).value = category
        logger.info(f'   - category {category};')

        tags_soup = soup.select_one('dt:-soup-contains("Tagi:")')
        tags = ", ".join(
            [tag.text.strip() for tag in tags_soup.find_next_sibling('dd').find_all('a')]) if tags_soup else ""
        sheet.cell(row=i + 1, column=10).value = tags

        description_text = description.get_text().replace('\n', '').strip() if description is not None else 'Brak opisu.'
        sheet.cell(row=i + 1, column=11).value = description_text
        logger.info(f"   - {description_text.split('.')[0][:100]};")

        logger.info(f'Completed {round((i - row_from + 1) * 100 / (row_to - row_from))}%.\n')

        i = 0
        while i < 3:
            try:
                wb.save(file_path)
                break
            except PermissionError:
                i += 1
                remaining_time = 15
                logger.warning("You don't have permission to access the file.")
                while remaining_time > 0:
                    logger.warning("Retry in %d sec.", remaining_time)
                    time.sleep(1)
                    remaining_time -= 1

    if counter == 1:
        logger.info(f'All {row_to - row_from} items have been found!')
    elif missing_items:
        logger.info(f'The following items were not found ({counter - 1} of {row_to - row_from}): \n - {chr(10).join(missing_items)};')
    else:
        logger.info(f'All {row_to - row_from} items have been found, but some have missing data.')


def file_check():
    if file_path.exists():
        main()
    else:
        filename = 'ebook.xlsx'
        wb = openpyxl.Workbook()
        wb.save(filename)
        main()


def option_1():
    print(f'In the active sheet {sheet.title} I write the information for {sheet.max_row - sheet.min_row} items.')
    row_from = int(sheet.min_row)
    row_to = int(sheet.max_row)
    main(row_from, row_to)


def option_2():
    print(f'The active sheet {sheet.title} has rows {sheet.min_row} through {sheet.max_row} filled in.')
    row_from = int(input('Row from: '))
    row_to = int(input('Row to: '))
    main(row_from, row_to)


def option_3():
    print(f'There are {sheet.max_row - sheet.min_row} items in the active sheet {sheet.title} (rows {sheet.min_row} to {sheet.max_row}).')
    print(f'The other sheets: {wb.sheetnames}')


def option_4():
    return True


options = {
    "1": option_1,
    "2": option_2,
    "3": option_3,
    "4": option_4
}

while True:
    print("+----------------------------------+")
    print("|        ...::: MENU :::...        |")
    print("+----------------------------------+")
    print("| 1. Find all items                |")
    print("| 2. Find selected items           |")
    print("| 3. Info                          |")
    print("| 4. Exit                          |")
    print("+----------------------------------+")

    option = input("Select an option: ")

    if option in options:
        if options[option]():
            break
    else:
        print("Incorrect entry, try again!")